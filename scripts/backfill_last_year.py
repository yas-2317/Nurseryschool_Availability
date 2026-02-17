#!/usr/bin/env python3
# -*- coding: utf-8 -*-

from __future__ import annotations

import csv
import io
import json
import os
import re
from dataclasses import dataclass
from datetime import date
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

import requests
from bs4 import BeautifulSoup
from openpyxl import load_workbook

CITY_PAGE = "https://www.city.yokohama.lg.jp/kosodate-kyoiku/hoiku-yoji/shisetsu/riyou/info/nyusho-jokyo.html"

WARD_FILTER = (os.getenv("WARD_FILTER", "") or "").strip() or None
MONTHS_BACK = int(os.getenv("MONTHS_BACK", "24"))
FORCE = (os.getenv("FORCE_BACKFILL", "0") == "1")

# master を backfill 時点で注入する（1推奨）
APPLY_MASTER = (os.getenv("APPLY_MASTER", "1") == "1")

ROOT = Path(__file__).resolve().parents[1]
DATA_DIR = ROOT / "data"
DATA_DIR.mkdir(parents=True, exist_ok=True)
MASTER_CSV = DATA_DIR / "master_facilities.csv"
MONTHS_JSON = DATA_DIR / "months.json"

REQ_HEADERS = {
    # GitHub Actions などで「UAなし=変なHTML」が返るのを避ける
    "User-Agent": "Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120 Safari/537.36",
    "Accept-Language": "ja,en-US;q=0.8,en;q=0.6",
    "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
}


# ---------- small utils ----------
def norm(s: Any) -> str:
    if s is None:
        return ""
    x = str(s).replace("　", " ")
    x = re.sub(r"\s+", "", x)
    return x.strip()


def safe(x: Any) -> str:
    return "" if x is None else str(x)


def to_int(x: Any) -> Optional[int]:
    if x is None:
        return None
    s = str(x).strip()
    if s == "" or s.lower() == "nan":
        return None
    if s in ("-", "－", "‐", "—", "―"):
        return 0
    try:
        return int(float(s))
    except Exception:
        return None


def sum_opt(*vals: Optional[int]) -> Optional[int]:
    xs = [v for v in vals if v is not None]
    return sum(xs) if xs else None


def ratio_opt(wait: Optional[int], cap: Optional[int]) -> Optional[float]:
    if wait is None or cap in (None, 0):
        return None
    return wait / cap


def month_floor(d: date) -> date:
    return date(d.year, d.month, 1)


def add_months(d: date, delta: int) -> date:
    y, m = d.year, d.month
    m2 = m + delta
    y += (m2 - 1) // 12
    m2 = (m2 - 1) % 12 + 1
    return date(y, m2, 1)


def iso(d: date) -> str:
    return d.isoformat()


def sanitize_header(header: List[str]) -> List[str]:
    out: List[str] = []
    seen: Dict[str, int] = {}
    for i, h in enumerate(header):
        h2 = (h or "").strip()
        if h2 == "":
            h2 = f"col{i}"
        if h2 in seen:
            seen[h2] += 1
            h2 = f"{h2}_{seen[h2]}"
        else:
            seen[h2] = 0
        out.append(h2)
    return out


def reiwa_to_ad(y_reiwa: int) -> int:
    # Reiwa 1 = 2019
    return 2018 + y_reiwa


# ---------- month detection (supports fiscal year hint) ----------
def extract_month_from_text(text: str, base_year_hint: Optional[int] = None) -> Optional[str]:
    """
    例:
      '【令和８年２月１日時点】' → 2026-02-01
      '2026年2月1日' → 2026-02-01
      '2月1日' / '４月１日' → base_year_hint から復元
         - 4〜12月 → base_year_hint
         - 1〜2月 → base_year_hint + 1
         - 3月は毎年公開無し想定（あっても base_year_hint+1 で復元）
    """
    if not text:
        return None

    t = str(text)
    z2h = str.maketrans("０１２３４５６７８９", "0123456789")
    t = t.translate(z2h)

    # Reiwa full date
    m = re.search(r"令和\s*([0-9]+)\s*年\s*([0-9]+)\s*月\s*1\s*日", t)
    if m:
        ry = int(m.group(1))
        mm = int(m.group(2))
        y = reiwa_to_ad(ry)
        return date(y, mm, 1).isoformat()

    # Gregorian full date
    m = re.search(r"([0-9]{4})\s*年\s*([0-9]{1,2})\s*月\s*1\s*日", t)
    if m:
        y = int(m.group(1))
        mm = int(m.group(2))
        return date(y, mm, 1).isoformat()

    # Month only (needs base_year_hint)
    m = re.search(r"([0-9]{1,2})\s*月\s*1\s*日", t)
    if m and base_year_hint:
        mm = int(m.group(1))
        y = base_year_hint if mm >= 4 else (base_year_hint + 1)
        return date(y, mm, 1).isoformat()

    return None


def detect_month_from_rows(rows: List[Dict[str, str]], base_year_hint: Optional[int] = None) -> Optional[str]:
    # 一部のExcelで「更新日」が入ってるケース
    if not rows:
        return None
    for k in ("更新日", "更新年月日", "更新日時", "更新年月"):
        v = str(rows[0].get(k, "")).strip()
        if v:
            v = v[:10].replace("/", "-")
            try:
                y, m, _ = v.split("-")
                return date(int(y), int(m), 1).isoformat()
            except Exception:
                return None

    # 行内に「2月1日」だけ出るケースもあるので拾えるなら拾う
    sample = " ".join(list(rows[0].values())[:10])
    m2 = extract_month_from_text(sample, base_year_hint=base_year_hint)
    return m2


# ---------- master apply ----------
def load_master() -> Dict[str, Dict[str, str]]:
    if not MASTER_CSV.exists():
        return {}
    out: Dict[str, Dict[str, str]] = {}
    with MASTER_CSV.open("r", encoding="utf-8-sig", newline="") as f:
        for row in csv.DictReader(f):
            fid = safe(row.get("facility_id")).strip()
            if fid:
                out[fid] = {k: safe(v) for k, v in row.items()}
    return out


def as_int_str(x: Any) -> Optional[str]:
    s = safe(x).strip()
    if s == "" or s.lower() == "null" or s == "-":
        return None
    try:
        return str(int(float(s)))
    except Exception:
        return None


def apply_master_to_facility(f: Dict[str, Any], m: Dict[str, str]) -> int:
    updated = 0
    mapping = {
        "address": "address",
        "lat": "lat",
        "lng": "lng",
        "map_url": "map_url",
        "facility_type": "facility_type",
        "phone": "phone",
        "website": "website",
        "notes": "notes",
        "nearest_station": "nearest_station",
        "name_kana": "name_kana",
        "station_kana": "station_kana",
    }
    for jkey, mkey in mapping.items():
        mv = safe(m.get(mkey)).strip()
        if mv == "":
            continue
        cur = safe(f.get(jkey)).strip()
        if cur != mv:
            f[jkey] = mv
            updated += 1

    wm = as_int_str(m.get("walk_minutes"))
    if wm is not None:
        cur = safe(f.get("walk_minutes")).strip()
        if cur != wm:
            f["walk_minutes"] = wm
            updated += 1

    return updated


# ---------- scraping ----------
@dataclass(frozen=True)
class ExcelLink:
    url: str
    base_year_hint: Optional[int]  # fiscal base year (e.g., 2024 for 令和6年度)


def _infer_base_year_from_context(context: str) -> Optional[int]:
    if not context:
        return None
    ctx = context
    z2h = str.maketrans("０１２３４５６７８９", "0123456789")
    ctx = ctx.translate(z2h)

    # 令和7年度 / 令和７年度
    m = re.search(r"令和\s*([0-9]+)\s*年度", ctx)
    if m:
        ry = int(m.group(1))
        return reiwa_to_ad(ry)

    # 2024年度
    m = re.search(r"([0-9]{4})\s*年度", ctx)
    if m:
        return int(m.group(1))

    return None


def scrape_excel_urls() -> Dict[str, List[ExcelLink]]:
    """
    横浜市ページから Excel リンク（.xls/.xlsx/.xlsm）を頑丈に拾って分類する。
    - 最新（単月更新の excel）も
    - 過去年度（令和7/6/5年度…の excel）も全部拾う
    - さらに「年度（base_year_hint）」を文脈から推定して持たせる
    """
    r = requests.get(CITY_PAGE, timeout=30, headers=REQ_HEADERS)
    r.raise_for_status()
    html = r.text

    # 空に近いHTMLが返るケースを早期検知
    if len(html) < 5000:
        raise RuntimeError(f"CITY_PAGE のHTMLが小さすぎます（len={len(html)}）。User-Agent対策が必要かもしれません。")

    soup = BeautifulSoup(html, "html.parser")

    # ページ全体から「今年度（令和7年度など）」が取れるなら default にする
    page_text = soup.get_text(" ", strip=True)
    default_base_year = _infer_base_year_from_context(page_text)

    found: List[ExcelLink] = []
    for a in soup.select("a[href]"):
        href = (a.get("href") or "").strip()
        if not href:
            continue

        href_abs = href if href.startswith("http") else requests.compat.urljoin(CITY_PAGE, href)
        ul = href_abs.lower()
        if not (ul.endswith(".xlsx") or ul.endswith(".xlsm") or ul.endswith(".xls") or ".xlsx?" in ul or ".xlsm?" in ul or ".xls?" in ul):
            continue

        # 近傍のテキストから年度を推定
        ctx_parts = []
        ctx_parts.append((a.get_text() or "").strip())
        if a.parent:
            ctx_parts.append(a.parent.get_text(" ", strip=True))
        if a.parent and a.parent.parent:
            ctx_parts.append(a.parent.parent.get_text(" ", strip=True))
        ctx = " ".join([p for p in ctx_parts if p])

        by = _infer_base_year_from_context(ctx) or default_base_year
        found.append(ExcelLink(url=href_abs, base_year_hint=by))

    # HTML直書きURLも拾う（念のため）
    for u in re.findall(r"https?://[^\s\"']+\.(?:xlsx|xlsm|xls)(?:\?[^\s\"']*)?", html, flags=re.I):
        found.append(ExcelLink(url=u, base_year_hint=default_base_year))

    # unique by url（base_year_hintは「より具体的なもの」を優先）
    uniq: Dict[str, ExcelLink] = {}
    for e in found:
        if e.url not in uniq:
            uniq[e.url] = e
        else:
            if uniq[e.url].base_year_hint is None and e.base_year_hint is not None:
                uniq[e.url] = e

    all_links = list(uniq.values())

    # classify
    urls: Dict[str, List[ExcelLink]] = {"accept": [], "wait": [], "enrolled": []}

    def kind_of(url: str) -> Optional[str]:
        ul = url.lower()
        # ファイル名で強いヒント
        if "0932_" in ul:
            return "accept"
        if "0933_" in ul or "0929_" in ul:
            return "wait"
        if "0934_" in ul or "0923_" in ul:
            return "enrolled"
        # 単語ヒント
        if "ukire" in ul or "受入" in ul:
            return "accept"
        if "mati" in ul or "待ち" in ul:
            return "wait"
        if "jido" in ul or "児童" in ul:
            return "enrolled"
        return None

    # まず url で分類
    for e in all_links:
        k = kind_of(e.url)
        if k:
            urls[k].append(e)

    # もし url だけで分類できないリンクがあるなら、リンク先は見に行かずに「アンカーテキスト」で補完…は難しいので、
    # ここでは「分類済みがゼロ」のときのみ「全部候補」として落とす（最低限落ちないため）
    # ※本来は a のテキストで分類が良いが、上で拾った found は e.url しか持たない（十分なはず）
    if not urls["accept"] and not urls["wait"] and not urls["enrolled"]:
        # ここまで来るのは「Excelリンク自体が拾えていない」ケース
        sample = [e.url for e in all_links[:10]]
        raise RuntimeError(f"Excelリンクが拾えません（accept=0, wait=0, sample={sample}）")

    # dedup per kind, keep order (newer first is not guaranteed; but ok)
    def dedup(lst: List[ExcelLink]) -> List[ExcelLink]:
        seen: Set[str] = set()
        out: List[ExcelLink] = []
        for e in lst:
            if e.url in seen:
                continue
            seen.add(e.url)
            out.append(e)
        return out

    for k in list(urls.keys()):
        urls[k] = dedup(urls[k])

    print("XLS links found:", {k: len(v) for k, v in urls.items()}, "default_base_year:", default_base_year)
    return urls


# ---------- Excel parsing ----------
def sheet_to_rows(ws) -> List[List[Any]]:
    rows: List[List[Any]] = []
    max_r = min(ws.max_row or 0, 6000)
    max_c = min(ws.max_column or 0, 120)
    for r in range(1, max_r + 1):
        row = []
        for c in range(1, max_c + 1):
            row.append(ws.cell(r, c).value)
        rows.append(row)
    return rows


def find_header_index(rows: List[List[Any]]) -> Optional[int]:
    keywords = ("施設", "合計", "0歳", "０歳", "1歳", "１歳", "受入", "待ち", "児童")
    best_i, best_score = None, -1
    for i, row in enumerate(rows[:120]):
        cells = ["" if v is None else str(v) for v in row]
        nonempty = sum(1 for c in cells if c.strip() != "")
        has_kw = any(any(k in c for k in keywords) for c in cells)
        score = nonempty + (10 if has_kw else 0)
        if nonempty >= 5 and score > best_score:
            best_i, best_score = i, score
    return best_i


def parse_sheet(ws, base_year_hint: Optional[int] = None) -> Tuple[Optional[str], List[Dict[str, str]]]:
    rows = sheet_to_rows(ws)

    month = extract_month_from_text(ws.title, base_year_hint=base_year_hint)
    if month is None:
        for r in rows[:20]:
            for v in r[:10]:
                month = extract_month_from_text("" if v is None else str(v), base_year_hint=base_year_hint)
                if month:
                    break
            if month:
                break

    hidx = find_header_index(rows)
    if hidx is None:
        return month, []

    header = sanitize_header([("" if v is None else str(v)) for v in rows[hidx]])
    out: List[Dict[str, str]] = []

    empty_streak = 0
    for r in rows[hidx + 1 :]:
        vals = [("" if v is None else str(v)) for v in r]
        if all(v.strip() == "" for v in vals):
            empty_streak += 1
            if empty_streak >= 10:
                break
            continue
        empty_streak = 0
        d = {header[i]: vals[i] if i < len(vals) else "" for i in range(len(header))}
        out.append(d)

    m2 = detect_month_from_rows(out, base_year_hint=base_year_hint)
    if m2:
        month = m2

    return month, out


def read_xlsx(url: str, base_year_hint: Optional[int]) -> Dict[str, List[Dict[str, str]]]:
    """
    xlsx 1ファイル → {month: rows} を返す
    base_year_hint は「年度（4月開始）の基準年」。シートに年が書いてない場合の補完に使う。
    """
    print("download:", url)
    r = requests.get(url, timeout=120, headers=REQ_HEADERS)
    r.raise_for_status()

    wb = load_workbook(io.BytesIO(r.content), data_only=True)

    mp: Dict[str, List[Dict[str, str]]] = {}
    for ws in wb.worksheets:
        month, rows = parse_sheet(ws, base_year_hint=base_year_hint)
        if month and rows:
            mp[month] = rows

    if mp:
        rng = (sorted(mp.keys())[0], sorted(mp.keys())[-1])
        print("  parsed months:", len(mp), "range:", rng, "base_year_hint:", base_year_hint)
    else:
        print("  parsed months: 0", "base_year_hint:", base_year_hint)
    return mp


# ---------- column guessing / metrics ----------
def guess_facility_id_key(rows: List[Dict[str, str]]) -> str:
    header = list(rows[0].keys())

    candidates = [
        "施設番号", "施設・事業所番号", "施設事業所番号", "事業所番号",
        "施設ID", "施設ＩＤ", "施設・事業所ID", "施設・事業所ＩＤ",
        "施設No", "施設Ｎｏ", "事業所No", "事業所Ｎｏ",
    ]
    for k in candidates:
        if k in rows[0]:
            return k

    patterns = ("番号", "ID", "ＩＤ", "No", "Ｎｏ", "NO", "ＮＯ")
    for k in header:
        if any(p in k for p in patterns) and ("施設" in k or "事業所" in k):
            return k

    N = min(200, len(rows))
    digit_re = re.compile(r"^\d{4,}$")
    best_key, best_score = None, -1
    for k in header:
        score = 0
        for i in range(N):
            v = str(rows[i].get(k, "")).strip()
            if digit_re.match(v):
                score += 1
        if score > best_score:
            best_key, best_score = k, score

    if best_key and best_score >= max(10, int(N * 0.30)):
        return best_key

    raise RuntimeError("施設番号列が見つかりません")


def index_by_key(rows: List[Dict[str, str]], key: str) -> Dict[str, Dict[str, str]]:
    out: Dict[str, Dict[str, str]] = {}
    for r in rows:
        v = str(r.get(key, "")).strip()
        if v:
            out[v] = r
    return out


def pick_ward_key(row: Dict[str, str]) -> Optional[str]:
    for k in ("施設所在区", "所在区", "区名"):
        if k in row:
            return k
    for k in row.keys():
        if "区" in k:
            return k
    return None


def pick_name_key(row: Dict[str, str]) -> Optional[str]:
    for k in ("施設名", "施設・事業名", "施設・事業所名", "事業名"):
        if k in row:
            return k
    for k in row.keys():
        if "施設" in k and "区" not in k:
            return k
    return None


def get_total(row: Dict[str, str]) -> Optional[int]:
    if not row:
        return None
    if "合計" in row and str(row.get("合計", "")).strip() != "":
        return to_int(row.get("合計"))
    for k in row.keys():
        if "合計" in k and str(row.get(k, "")).strip() != "":
            return to_int(row.get(k))
    return None


def get_age_value(row: Dict[str, str], age: int) -> Optional[int]:
    if not row:
        return None
    z = "０１２３４５"
    pats = [f"{age}歳児", f"{age}歳", z[age] + "歳児", z[age] + "歳"]
    for p in pats:
        if p in row and str(row.get(p, "")).strip() != "":
            return to_int(row.get(p))
    for k in row.keys():
        if any(p in k for p in pats) and str(row.get(k, "")).strip() != "":
            return to_int(row.get(k))
    return None


def build_age_groups(ar: Dict[str, str], wr: Dict[str, str], er: Dict[str, str]) -> Tuple[Dict[str, Any], Dict[str, Any]]:
    ages_0_5: Dict[str, Dict[str, Any]] = {}
    for i in range(6):
        a = get_age_value(ar, i)
        w = get_age_value(wr, i) if wr else None
        e = get_age_value(er, i) if er else None
        cap_est = (e + a) if (e is not None and a is not None) else None
        ages_0_5[str(i)] = {
            "accept": a,
            "wait": w,
            "enrolled": e,
            "capacity_est": cap_est,
            "wait_per_capacity_est": ratio_opt(w, cap_est),
        }

    g0, g1, g2 = ages_0_5["0"], ages_0_5["1"], ages_0_5["2"]
    g3, g4, g5 = ages_0_5["3"], ages_0_5["4"], ages_0_5["5"]

    w_35 = sum_opt(g3.get("wait"), g4.get("wait"), g5.get("wait"))
    cap_35 = sum_opt(g3.get("capacity_est"), g4.get("capacity_est"), g5.get("capacity_est"))

    age_groups = {
        "0": g0,
        "1": g1,
        "2": g2,
        "3-5": {
            "accept": sum_opt(g3.get("accept"), g4.get("accept"), g5.get("accept")),
            "wait": w_35,
            "enrolled": sum_opt(g3.get("enrolled"), g4.get("enrolled"), g5.get("enrolled")),
            "capacity_est": cap_35,
            "wait_per_capacity_est": ratio_opt(w_35, cap_35),
        },
    }
    return age_groups, ages_0_5


# ---------- main backfill ----------
def main() -> None:
    print("BACKFILL start. ward=", WARD_FILTER, "months_back=", MONTHS_BACK, "force=", FORCE, "apply_master=", APPLY_MASTER)

    urls = scrape_excel_urls()
    master = load_master() if APPLY_MASTER else {}
    target = norm(WARD_FILTER) if WARD_FILTER else None

    acc_by_month: Dict[str, List[Dict[str, str]]] = {}
    wai_by_month: Dict[str, List[Dict[str, str]]] = {}
    enr_by_month: Dict[str, List[Dict[str, str]]] = {}

    # accept/wait/enrolled それぞれ「最新+過去年度」のExcelが入る前提
    for e in urls["accept"]:
        try:
            acc_by_month.update(read_xlsx(e.url, e.base_year_hint))
        except Exception as ex:
            print("WARN accept xlsx failed:", e.url, ex)

    for e in urls["wait"]:
        try:
            wai_by_month.update(read_xlsx(e.url, e.base_year_hint))
        except Exception as ex:
            print("WARN wait xlsx failed:", e.url, ex)

    for e in urls["enrolled"]:
        try:
            enr_by_month.update(read_xlsx(e.url, e.base_year_hint))
        except Exception as ex:
            print("WARN enrolled xlsx failed:", e.url, ex)

    if not acc_by_month:
        raise RuntimeError("受入可能数の月次が1つも取れませんでした")

    end = month_floor(date.today())
    start = add_months(end, -(MONTHS_BACK - 1))
    want: List[str] = []
    cur = start
    while cur <= end:
        want.append(iso(cur))
        cur = add_months(cur, 1)

    available = [m for m in want if m in acc_by_month]
    missing = [m for m in want if m not in acc_by_month]
    print("want months:", len(want), "available:", len(available), "missing:", missing[:30], ("..." if len(missing) > 30 else ""))

    existing_months: List[str] = []
    if MONTHS_JSON.exists():
        try:
            existing_months = json.loads(MONTHS_JSON.read_text(encoding="utf-8")).get("months", [])
        except Exception:
            existing_months = []

    changed_any = 0

    for m in available:
        out_path = DATA_DIR / f"{m}.json"
        if out_path.exists() and not FORCE:
            print("skip exists:", out_path.name)
            continue

        accept_rows = acc_by_month.get(m, [])
        wait_rows = wai_by_month.get(m, [])
        enrolled_rows = enr_by_month.get(m, [])

        if not accept_rows:
            continue

        fid_a = guess_facility_id_key(accept_rows)
        A = index_by_key(accept_rows, fid_a)

        W = {}
        if wait_rows:
            try:
                fid_w = guess_facility_id_key(wait_rows)
                W = index_by_key(wait_rows, fid_w)
            except Exception:
                W = {}

        E = {}
        if enrolled_rows:
            try:
                fid_e = guess_facility_id_key(enrolled_rows)
                E = index_by_key(enrolled_rows, fid_e)
            except Exception:
                E = {}

        ward_key = pick_ward_key(accept_rows[0]) if accept_rows else None
        name_key = pick_name_key(accept_rows[0]) if accept_rows else None

        facilities: List[Dict[str, Any]] = []
        injected_cells = 0

        for fid, ar in A.items():
            ward = norm(ar.get(ward_key)) if ward_key else ""
            ward = ward.replace("横浜市", "")
            if target and target not in ward:
                continue

            wr = W.get(fid, {})
            er = E.get(fid, {})

            name = str(ar.get(name_key, "")).strip() if name_key else ""

            tot_accept = get_total(ar)
            tot_wait = get_total(wr) if wr else None
            tot_enrolled = get_total(er) if er else None
            cap_est = (tot_enrolled + tot_accept) if (tot_enrolled is not None and tot_accept is not None) else None

            age_groups, ages_0_5 = build_age_groups(ar, wr, er)

            fobj: Dict[str, Any] = {
                "id": fid,
                "name": name,
                "name_kana": "",
                "ward": ward,
                "address": "",
                "lat": "",
                "lng": "",
                "map_url": "",
                "facility_type": "",
                "phone": "",
                "website": "",
                "notes": "",
                "nearest_station": "",
                "station_kana": "",
                "walk_minutes": None,
                "updated": m,
                "totals": {
                    "accept": tot_accept,
                    "wait": tot_wait,
                    "enrolled": tot_enrolled,
                    "capacity_est": cap_est,
                    "wait_per_capacity_est": ratio_opt(tot_wait, cap_est),
                },
                "age_groups": age_groups,
                "ages_0_5": ages_0_5,
            }

            if APPLY_MASTER:
                mm = master.get(fid)
                if mm:
                    injected_cells += apply_master_to_facility(fobj, mm)

            facilities.append(fobj)

        out = {"month": m, "ward": (WARD_FILTER or "横浜市"), "facilities": facilities}
        out_path.write_text(json.dumps(out, ensure_ascii=False, indent=2), encoding="utf-8")
        print("wrote:", out_path.name, "facilities:", len(facilities), "master_injected_cells:", injected_cells)
        changed_any += 1

    ms = set(existing_months)
    for m in available:
        p = DATA_DIR / f"{m}.json"
        if p.exists() and p.stat().st_size > 200:
            ms.add(m)
    MONTHS_JSON.write_text(json.dumps({"months": sorted(ms)}, ensure_ascii=False, indent=2), encoding="utf-8")
    print("updated months.json:", len(ms), "changed_month_files:", changed_any)


if __name__ == "__main__":
    main()
